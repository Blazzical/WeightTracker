"""Import daily log data from the weightloss 2026.xlsx spreadsheet."""

import re
import openpyxl
import models
from seed_data import seed


def parse_kj_cal_from_text(text):
    """Extract kJ and cal values from text like 'Cappuccino (628 kJ / 150 cal)'."""
    m = re.search(r'\((\d+)\s*kJ\s*/\s*(\d+)\s*cal\)', text)
    if m:
        return float(m.group(1)), float(m.group(2))
    m = re.search(r'\((\d+)\s*kJ,\s*(\d+)\s*cal\)', text)
    if m:
        return float(m.group(1)), float(m.group(2))
    return None, None


def strip_nutrition_info(text):
    """Remove parenthetical nutrition info from text."""
    return re.sub(r'\s*\(\d+\s*kJ\s*[/,]\s*\d+\s*cal\)', '', text).strip()


def find_meal_by_name(name, all_meals):
    """Try to find a matching meal (case-insensitive, fuzzy)."""
    name_lower = name.lower().strip()

    # Direct match
    for m in all_meals:
        if m['name'].lower() == name_lower:
            return m

    # Normalize common variations
    normalized = name_lower
    normalized = normalized.replace('w/ half water', 'half water')
    normalized = normalized.replace('(w/ half water)', '(half water)')
    normalized = normalized.replace('black berries', 'blackberries')
    normalized = normalized.replace('baked beans', 'beans')

    for m in all_meals:
        m_norm = m['name'].lower()
        m_norm = m_norm.replace('w/ half water', 'half water')
        m_norm = m_norm.replace('(w/ half water)', '(half water)')
        if m_norm == normalized:
            return m

    # Substring match (meal name contained in description or vice versa)
    for m in all_meals:
        m_lower = m['name'].lower()
        if m_lower in name_lower or name_lower in m_lower:
            return m

    return None


def find_component_by_name(name, all_components):
    """Try to find a matching component."""
    name_lower = name.lower().strip()
    for c in all_components:
        if c['name'].lower() == name_lower:
            return c
    # Substring match
    for c in all_components:
        if c['name'].lower() in name_lower or name_lower in c['name'].lower():
            return c
    return None


def get_or_create_meal_for_entry(food_text, all_meals):
    """Find or create a meal for a food entry. Returns meal_id."""
    clean_name = strip_nutrition_info(food_text)
    kj, cal = parse_kj_cal_from_text(food_text)

    if not clean_name or clean_name == '-':
        return None

    # Try to find existing meal
    meal = find_meal_by_name(clean_name, all_meals)
    if meal:
        return meal['id']

    # Create new meal with override values if we have them
    meal_id = models.add_meal(clean_name, kj_override=kj, calories_override=cal)
    if meal_id:
        # Refresh the meals list
        new_meal = models.get_meal(meal_id)
        if new_meal:
            all_meals.append(new_meal)
        return meal_id
    else:
        # Name collision - find the existing one
        meal = find_meal_by_name(clean_name, all_meals)
        return meal['id'] if meal else None


def import_spreadsheet():
    """Import all daily data from the spreadsheet."""
    models.init_db()
    seed()

    wb = openpyxl.load_workbook(r'E:\weightloss 2026.xlsx', data_only=True)
    ws = wb['Sheet1']

    all_meals = models.get_all_meals()
    all_components = models.get_all_components()

    stats = {'logs_created': 0, 'logs_updated': 0, 'entries_added': 0, 'meals_created': 0, 'skipped': 0}

    for row_num in range(2, ws.max_row + 1):
        date_val = ws.cell(row_num, 1).value  # Column A
        if not date_val:
            continue

        # Convert date
        if hasattr(date_val, 'strftime'):
            date_str = date_val.strftime('%Y-%m-%d')
        else:
            continue

        weight_morning = ws.cell(row_num, 2).value   # Column B
        exercise = ws.cell(row_num, 3).value or ''    # Column C
        breakfast = ws.cell(row_num, 4).value          # Column D
        lunch = ws.cell(row_num, 5).value              # Column E
        dinner = ws.cell(row_num, 6).value             # Column F
        extras = ws.cell(row_num, 7).value             # Column G
        coffees = ws.cell(row_num, 8).value            # Column H
        weight_night = ws.cell(row_num, 10).value      # Column J

        # Check if any food data exists for this day
        has_food = any([breakfast, lunch, dinner, extras, coffees])
        if not has_food and not weight_morning and not weight_night:
            continue

        # Create/get daily log
        existing_log = models.get_daily_log(date_str)
        if existing_log:
            # Update weight/exercise but check if entries already exist
            existing_entries = models.get_daily_entries(existing_log['id'])
            if existing_entries:
                # Don't add duplicate entries - just update the log metadata
                wm = weight_morning if weight_morning else existing_log.get('weight_morning')
                wn = weight_night if weight_night else existing_log.get('weight_night')
                ex = exercise if exercise else existing_log.get('exercise', '')
                models.update_daily_log(existing_log['id'], wm, wn, ex)
                print(f"  {date_str}: already has {len(existing_entries)} entries, updated weight/exercise only")
                stats['logs_updated'] += 1
                stats['skipped'] += 1
                continue
            log = existing_log
        else:
            log = models.get_or_create_daily_log(date_str)
            stats['logs_created'] += 1

        # Update weight and exercise
        models.update_daily_log(log['id'], weight_morning, weight_night, str(exercise))

        # Add food entries
        meal_time_map = [
            (breakfast, 'breakfast'),
            (lunch, 'lunch'),
            (dinner, 'dinner'),
        ]

        for food_text, meal_time in meal_time_map:
            if not food_text or food_text.strip() == '-':
                continue

            meal_id = get_or_create_meal_for_entry(str(food_text), all_meals)
            if meal_id:
                models.add_daily_entry(log['id'], meal_id=meal_id, quantity=1, meal_time=meal_time)
                stats['entries_added'] += 1
                print(f"  {date_str} [{meal_time}]: {strip_nutrition_info(str(food_text))}")

        # Handle extras (may have multiple items separated by commas or periods)
        if extras and str(extras).strip() != '-':
            extras_str = str(extras)
            # Split on commas that aren't inside parentheses
            parts = re.split(r',\s*(?![^(]*\))', extras_str)
            for part in parts:
                part = part.strip()
                if not part or part == '-':
                    continue
                # Split on '. ' for sentence-separated items
                sub_parts = [p.strip() for p in re.split(r'\.\s+', part) if p.strip()]
                for sub in sub_parts:
                    if not sub or sub == '-':
                        continue
                    meal_id = get_or_create_meal_for_entry(sub, all_meals)
                    if meal_id:
                        models.add_daily_entry(log['id'], meal_id=meal_id, quantity=1, meal_time='snack')
                        stats['entries_added'] += 1
                        print(f"  {date_str} [snack]: {strip_nutrition_info(sub)}")

        # Handle coffees
        if coffees:
            try:
                coffee_cal = float(coffees)
                if coffee_cal > 0:
                    # Find Coffee w/ 50mL Milk (42 cal each) or Cappuccino (150 cal)
                    coffee_meal = find_meal_by_name('Coffee w/ 50mL Milk', all_meals)
                    if coffee_meal and coffee_meal['total_calories'] > 0:
                        num_coffees = round(coffee_cal / coffee_meal['total_calories'])
                        if num_coffees > 0:
                            models.add_daily_entry(log['id'], meal_id=coffee_meal['id'],
                                                   quantity=num_coffees, meal_time='coffee')
                            stats['entries_added'] += 1
                            print(f"  {date_str} [coffee]: {num_coffees}x Coffee w/ 50mL Milk")
            except (ValueError, TypeError):
                pass

    # Refresh meals list to count new ones
    final_meals = models.get_all_meals()
    stats['meals_created'] = len(final_meals) - 16  # 16 were seeded

    print(f"\n--- Import complete ---")
    print(f"Daily logs created: {stats['logs_created']}")
    print(f"Daily logs updated: {stats['logs_updated']}")
    print(f"Daily entries added: {stats['entries_added']}")
    print(f"New meals created: {stats['meals_created']}")
    print(f"Days skipped (already had entries): {stats['skipped']}")


if __name__ == '__main__':
    import_spreadsheet()
